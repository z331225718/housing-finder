# Import/Export router for Excel file handling
from io import BytesIO
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl import Workbook
from fastapi import APIRouter, File, HTTPException, UploadFile, Depends
from fastapi.responses import StreamingResponse

from app.auth import get_current_user
from app.models import User, Community, Property
from app.database import get_db
from app import crud
from sqlalchemy.orm import Session

router = APIRouter(prefix="/import-export", tags=["import-export"])


# ============ Excel Template Generation ============

def generate_community_template() -> bytes:
    """Generate Excel template for community data entry."""
    wb = Workbook()
    ws = wb.active
    ws.title = "小区信息"

    # Headers
    headers = [
        "小区名称*", "所属区*", "详细地址", "物业费", "停车位", "建成年份",
        "周边配套/地铁", "对口小学", "对口中学", "环境打分(1-10)", "备注"
    ]
    ws.append(headers)

    # Example row
    example = [
        "示例小区", "浦东新区", "XX路123号", "2.5元/平/月", "地上50个,地下100个",
        2015, "地铁9号线, 商场", "明珠小学", "明珠中学", 8, "小区环境好"
    ]
    ws.append(example)

    # Set column widths
    ws.column_dimensions['A'].width = 15  # name
    ws.column_dimensions['B'].width = 12  # district
    ws.column_dimensions['C'].width = 25  # address
    ws.column_dimensions['D'].width = 15  # property_fee
    ws.column_dimensions['E'].width = 20  # parking
    ws.column_dimensions['F'].width = 10  # build_year
    ws.column_dimensions['G'].width = 20  # metro
    ws.column_dimensions['H'].width = 15  # primary_school
    ws.column_dimensions['I'].width = 15  # middle_school
    ws.column_dimensions['J'].width = 12  # environment_score
    ws.column_dimensions['K'].width = 30  # notes

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_property_template() -> bytes:
    """Generate Excel template for property data entry."""
    wb = Workbook()
    ws = wb.active
    ws.title = "房源信息"

    # Headers
    headers = [
        "小区名称*", "楼号", "单元", "房号", "面积(㎡)*", "户型",
        "楼层", "朝向", "装修情况", "挂牌价格(万)*", "租金(元/月)",
        "预计价格(万)", "看房日期(YYYY-MM-DD)", "备注"
    ]
    ws.append(headers)

    # Example row
    example = [
        "示例小区", "1", "1", "101", 120, "3室2厅",
        "中楼层", "南", "精装", 800, 6000,
        750, "2024-01-15", "采光好"
    ]
    ws.append(example)

    # Set column widths
    ws.column_dimensions['A'].width = 15  # community_name
    ws.column_dimensions['B'].width = 8   # building
    ws.column_dimensions['C'].width = 8   # unit
    ws.column_dimensions['D'].width = 8   # room
    ws.column_dimensions['E'].width = 12  # area
    ws.column_dimensions['F'].width = 12  # layout
    ws.column_dimensions['G'].width = 10  # floor
    ws.column_dimensions['H'].width = 10  # orientation
    ws.column_dimensions['I'].width = 12  # decoration
    ws.column_dimensions['J'].width = 15  # price
    ws.column_dimensions['K'].width = 15  # rent
    ws.column_dimensions['L'].width = 15  # expected_price
    ws.column_dimensions['M'].width = 18  # visit_date
    ws.column_dimensions['N'].width = 30  # notes

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


@router.get("/template/community")
async def download_community_template():
    """Download community Excel template."""
    content = generate_community_template()
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=community_template.xlsx"}
    )


@router.get("/template/property")
async def download_property_template():
    """Download property Excel template."""
    content = generate_property_template()
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=property_template.xlsx"}
    )


# ============ Excel Import ============

def parse_community_row(row: list) -> Dict[str, Any]:
    """Parse a community row from Excel."""
    return {
        "name": str(row[0]).strip() if row[0] else "",
        "district": str(row[1]).strip() if row[1] else "",
        "address": str(row[2]).strip() if len(row) > 2 and row[2] else "",
        "property_fee": str(row[3]).strip() if len(row) > 3 and row[3] else "",
        "parking": str(row[4]).strip() if len(row) > 4 and row[4] else "",
        "build_year": int(row[5]) if len(row) > 5 and row[5] and str(row[5]).isdigit() else None,
        "metro": str(row[6]).strip() if len(row) > 6 and row[6] else "",
        "primary_school": str(row[7]).strip() if len(row) > 7 and row[7] else "",
        "middle_school": str(row[8]).strip() if len(row) > 8 and row[8] else "",
        "environment_score": int(row[9]) if len(row) > 9 and row[9] and str(row[9]).isdigit() else None,
        "notes": str(row[10]).strip() if len(row) > 10 and row[10] else "",
    }


def parse_property_row(row: list, community_map: Dict[str, int]) -> Optional[Dict[str, Any]]:
    """Parse a property row from Excel."""
    community_name = str(row[0]).strip() if row[0] else ""

    if community_name not in community_map:
        return None

    # Parse visit_date
    visit_date = None
    if len(row) > 12 and row[12]:
        try:
            from datetime import datetime
            visit_date = datetime.strptime(str(row[12]).strip(), "%Y-%m-%d")
        except ValueError:
            pass

    return {
        "community_id": community_map[community_name],
        "building": str(row[1]).strip() if len(row) > 1 and row[1] else "",
        "unit": str(row[2]).strip() if len(row) > 2 and row[2] else "",
        "room": str(row[3]).strip() if len(row) > 3 and row[3] else "",
        "area": float(row[4]) if len(row) > 4 and row[4] else None,
        "layout": str(row[5]).strip() if len(row) > 5 and row[5] else "",
        "floor": str(row[6]).strip() if len(row) > 6 and row[6] else "",
        "orientation": str(row[7]).strip() if len(row) > 7 and row[7] else "",
        "decoration": str(row[8]).strip() if len(row) > 8 and row[8] else "",
        "price": float(row[9]) if len(row) > 9 and row[9] else None,
        "rent": float(row[10]) if len(row) > 10 and row[10] else None,
        "expected_price": float(row[11]) if len(row) > 11 and row[11] else None,
        "visit_date": visit_date,
        "notes": str(row[12]).strip() if len(row) > 13 and row[13] else "",
    }


@router.post("/community")
async def import_communities(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> dict:
    """Import communities from Excel file."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active

        # Skip header row
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        imported = []
        errors = []

        for idx, row in enumerate(rows):
            if not row or not row[0]:  # Skip empty rows
                continue

            try:
                data = parse_community_row(row)

                # Validate required fields
                if not data["name"]:
                    errors.append(f"Row {idx + 2}: 小区名称不能为空")
                    continue
                if not data["district"]:
                    errors.append(f"Row {idx + 2}: 所属区不能为空")
                    continue

                # Create community
                from app.schemas import CommunityCreate
                community = crud.create_community(db, CommunityCreate(**data))
                imported.append({"id": community.id, "name": community.name})

            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        return {
            "success": True,
            "imported": len(imported),
            "details": imported,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")


@router.post("/property")
async def import_properties(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> dict:
    """Import properties from Excel file."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active

        # First, get all existing communities to build name->id map
        communities = db.query(Community).all()
        community_map = {c.name: c.id for c in communities}

        # Skip header row
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        imported = []
        errors = []

        for idx, row in enumerate(rows):
            if not row or not row[0]:  # Skip empty rows
                continue

            try:
                data = parse_property_row(row, community_map)

                if not data:
                    errors.append(f"Row {idx + 2}: 小区 '{row[0]}' 不存在，请先创建小区")
                    continue

                # Validate required fields
                if not data["area"]:
                    errors.append(f"Row {idx + 2}: 面积不能为空")
                    continue
                if not data["price"]:
                    errors.append(f"Row {idx + 2}: 挂牌价格不能为空")
                    continue

                # Create property
                from app.schemas import PropertyCreate
                property_obj = crud.create_property(db, PropertyCreate(**data))
                imported.append({"id": property_obj.id, "area": property_obj.area, "price": property_obj.price})

            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        return {
            "success": True,
            "imported": len(imported),
            "details": imported,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
